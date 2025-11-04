from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import json
import re

class NutritionScraperComplete:
    def __init__(self, testing_mode=False):
        """Initialize the scraper with Chrome options
        
        Args:
            testing_mode (bool): If True, limits scraping for faster testing
        """
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 15)
        self.base_url = "https://eatsmart.housing.illinois.edu"
        self.testing_mode = testing_mode
        self.max_items_per_meal = 5 if testing_mode else None
    
    def determine_if_entree(self, item_name, serving_size=None, calories=None):
        """Determine if a menu item is an entrée based on various criteria"""
        
        # Convert to lowercase for comparison
        name_lower = item_name.lower() if item_name else ""
        
        # Lists of keywords that typically indicate entrées
        entree_keywords = [
            # Proteins
            'chicken', 'beef', 'pork', 'fish', 'salmon', 'tuna', 'tilapia', 'cod',
            'turkey', 'lamb', 'shrimp', 'steak', 'roast', 'grilled', 'baked',
            'fried', 'breaded', 'cutlet', 'fillet', 'filet', 'wings', 'tenders',
            'nuggets', 'patty', 'burger', 'meatball', 'sausage', 'bacon',
            
            # Main dishes
            'pasta', 'spaghetti', 'lasagna', 'ravioli', 'fettuccine', 'penne',
            'pizza', 'sandwich', 'wrap', 'burrito', 'taco', 'quesadilla',
            'stir fry', 'stir-fry', 'curry', 'masala', 'korma', 'tikka',
            'bowl', 'plate', 'entree', 'entrée', 'meal', 'dinner',
            
            # Asian dishes
            'teriyaki', 'kung pao', 'sweet and sour', 'lo mein', 'chow mein',
            'pad thai', 'fried rice', 'bibimbap', 'ramen', 'pho',
            
            # Other mains
            'omelet', 'omelette', 'scramble', 'benedict', 'frittata',
            'casserole', 'pot pie', 'stuffed', 'platter'
        ]
        
        # Keywords that typically indicate sides/non-entrées
        non_entree_keywords = [
            # Sides
            'rice', 'plain', 'white rice', 'brown rice', 'steamed rice',
            'mashed potato', 'french fries', 'fries', 'tots', 'hash brown',
            'vegetable', 'veggie', 'broccoli', 'carrot', 'corn', 'peas', 
            'green beans', 'asparagus', 'spinach', 'kale', 'lettuce',
            
            # Condiments and small items
            'sauce', 'gravy', 'dressing', 'dip', 'salsa', 'ketchup', 'mayo',
            'mayonnaise', 'mustard', 'butter', 'cream', 'cheese', 'sour cream',
            'syrup', 'honey', 'jam', 'jelly', 'sugar', 'salt', 'pepper',
            'oil', 'vinegar', 'spice', 'seasoning', 'topping',
            
            # Soups (unless it's a main soup)
            'broth', 'stock',
            
            # Breakfast sides
            'toast', 'muffin', 'bagel', 'donut', 'pastry', 'croissant',
            'cereal', 'oatmeal', 'porridge', 'granola', 'yogurt',
            
            # Snacks and desserts
            'cookie', 'cake', 'pie', 'brownie', 'dessert', 'ice cream',
            'pudding', 'jello', 'fruit', 'apple', 'banana', 'orange',
            
            # Beverages
            'juice', 'milk', 'coffee', 'tea', 'soda', 'water', 'drink',
            
            # Small items
            'crouton', 'cracker', 'chip', 'nuts', 'seeds',
            'garnish', 'cilantro', 'parsley', 'onion', 'garlic', 'ginger'
        ]
        
        # Special cases - soups can be entrées
        soup_entree_keywords = ['chowder', 'stew', 'chili', 'gumbo', 'bisque', 'congee']
        
        # Check if it's a soup that could be an entrée
        if 'soup' in name_lower:
            for keyword in soup_entree_keywords:
                if keyword in name_lower:
                    return True
            # Regular soup without protein is usually not an entrée
            has_protein = any(protein in name_lower for protein in 
                            ['chicken', 'beef', 'pork', 'seafood', 'shrimp', 'tofu', 'bean', 'lentil'])
            if has_protein:
                return True
            return False
        
        # Check for non-entrée keywords first (they're more specific)
        for keyword in non_entree_keywords:
            if keyword in name_lower:
                # But check for exceptions (e.g., "chicken rice bowl" should be entrée)
                has_entree_keyword = any(ek in name_lower for ek in entree_keywords)
                if has_entree_keyword:
                    # Item has both entrée and non-entrée keywords
                    # Use additional logic
                    if any(main in name_lower for main in ['bowl', 'platter', 'plate', 'meal']):
                        return True
                    # Check calorie count if available
                    if calories:
                        try:
                            cal_value = int(re.search(r'\d+', str(calories)).group())
                            if cal_value >= 200:  # High calorie items more likely to be entrées
                                return True
                        except:
                            pass
                return False
        
        # Check for entrée keywords
        for keyword in entree_keywords:
            if keyword in name_lower:
                return True
        
        # Additional logic based on serving size
        if serving_size and isinstance(serving_size, str):
            serving_lower = serving_size.lower()
            # Large serving sizes often indicate entrées
            if any(unit in serving_lower for unit in ['bowl', 'plate', 'platter', 'portion', 'serving']):
                return True
        
        # Check calorie count as last resort
        if calories and calories != 'N/A':
            try:
                cal_value = int(re.search(r'\d+', str(calories)).group())
                # Items with 250+ calories are more likely to be entrées
                if cal_value >= 250:
                    return True
            except:
                pass
        
        # Default to False if no clear indication
        return False
    
    def scrape_dining_structure(self):
        """Scrape all dining halls and their services from the dropdown menu"""
        try:
            print("Loading main page to extract dining hall structure...")
            self.driver.get(self.base_url + "/NetNutrition/1")
            time.sleep(4)
            
            print("Extracting dining halls and services from navigation dropdown...")
            
            dropdown = self.driver.find_element(By.ID, "nav-unit-selector")
            dropdown_items = dropdown.find_elements(By.CSS_SELECTOR, ".dropdown-item")
            
            dining_halls = []
            current_hall = None
            
            for item in dropdown_items:
                try:
                    link = item.find_element(By.TAG_NAME, "a")
                    name = link.get_attribute('title') or link.text.strip()
                    unit_id = link.get_attribute('data-unitoid')
                    
                    if not name or not unit_id or unit_id == '-1':
                        continue
                    
                    is_primary = 'text-primary' in link.get_attribute('class')
                    
                    if is_primary:
                        if current_hall and current_hall['dining_services']:
                            dining_halls.append(current_hall)
                        
                        current_hall = {
                            'dining_hall': name,
                            'unit_id': unit_id,
                            'dining_services': []
                        }
                    else:
                        if current_hall:
                            service = {
                                'service_name': name,
                                'service_id': unit_id
                            }
                            current_hall['dining_services'].append(service)
                
                except Exception as e:
                    continue
            
            if current_hall and current_hall['dining_services']:
                dining_halls.append(current_hall)
            
            return dining_halls
            
        except Exception as e:
            print(f"Error scraping dining structure: {str(e)}")
            return []
    
    def navigate_to_service(self, unit_id, service_name):
        """Navigate to a specific dining service"""
        try:
            print(f"\nNavigating to {service_name} (ID: {unit_id})...")
            
            self.driver.get(f"{self.base_url}/NetNutrition/1")
            time.sleep(4)
            
            dropdown = self.driver.find_element(By.ID, "nav-unit-selector")
            service_link = dropdown.find_element(By.CSS_SELECTOR, f"a[data-unitoid='{unit_id}']")
            
            self.driver.execute_script("arguments[0].click();", service_link)
            time.sleep(5)
            
            print("Service loaded")
            return True
            
        except Exception as e:
            print(f"Error: {str(e)}")
            return False
    
    def get_all_meals_structured(self):
        """Get all available meals organized by date and meal period"""
        try:
            print("Extracting meal structure...")
            time.sleep(3)
            
            results_panel = self.driver.find_element(By.ID, "navBarResults")
            
            # First, get the card structure that contains date/meal headers
            try:
                card = results_panel.find_element(By.CSS_SELECTOR, ".card")
                
                # Find all date/meal headers (format: "Tuesday, November 4, 2025-Lunch")
                date_meal_headers = card.find_elements(By.CSS_SELECTOR, ".card-header.bg-primary, .d-flex.card-header.bg-primary")
                
                print(f"Found {len(date_meal_headers)} date/meal headers")
                
                # Extract date and meal from each header
                date_meal_pairs = []
                for header in date_meal_headers:
                    header_text = header.text.strip()
                    if header_text and '-' in header_text:
                        # Split by the last hyphen to separate date and meal
                        # Format: "Tuesday, November 4, 2025-Lunch"
                        parts = header_text.rsplit('-', 1)
                        if len(parts) == 2:
                            date_part = parts[0].strip()
                            meal_part = parts[1].strip()
                            
                            # Clean up meal type (remove any extra text)
                            meal_types = ['Breakfast', 'Lunch', 'Dinner', 'Brunch', 'Late Night']
                            meal_type = None
                            for meal in meal_types:
                                if meal.lower() in meal_part.lower():
                                    meal_type = meal
                                    break
                            
                            if meal_type:
                                date_meal_pairs.append({
                                    'date': date_part,
                                    'meal_type': meal_type,
                                    'header_text': header_text
                                })
                                print(f"  Parsed: {date_part} - {meal_type}")
            except Exception as e:
                print(f"Could not find card structure, trying alternative approach: {str(e)}")
                date_meal_pairs = []
            
            # Get all menu items (the clickable li elements)
            menu_items = results_panel.find_elements(By.CSS_SELECTOR, "li.list-group-item")
            print(f"Found {len(menu_items)} menu items")
            
            structured_meals = []
            
            # Match menu items with their corresponding date/meal
            # If we have the same number of items as date/meal pairs, they likely correspond 1:1
            if len(date_meal_pairs) == len(menu_items):
                for i, menu_item in enumerate(menu_items):
                    if i < len(date_meal_pairs):
                        meal_info = {
                            'element': menu_item,
                            'date': date_meal_pairs[i]['date'],
                            'meal_type': date_meal_pairs[i]['meal_type'],
                            'onclick': menu_item.get_attribute('onclick')
                        }
                        structured_meals.append(meal_info)
            else:
                # Fallback: try to extract any info from the menu items themselves
                print("Number of headers doesn't match menu items, using fallback method")
                
                # Look for any text in the navBarResults that might contain date/meal info
                all_text = results_panel.text
                lines = all_text.split('\n')
                
                current_date = None
                current_meal = None
                
                for line in lines:
                    line = line.strip()
                    # Check if line contains a date
                    if any(day in line for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']):
                        if '2025' in line or '2024' in line:
                            # This is likely a date/meal line
                            if '-' in line:
                                parts = line.rsplit('-', 1)
                                if len(parts) == 2:
                                    current_date = parts[0].strip()
                                    meal_part = parts[1].strip()
                                    
                                    meal_types = ['Breakfast', 'Lunch', 'Dinner', 'Brunch', 'Late Night']
                                    for meal in meal_types:
                                        if meal.lower() in meal_part.lower():
                                            current_meal = meal
                                            break
                
                # Apply the last found date/meal to all menu items (temporary solution)
                for menu_item in menu_items:
                    meal_info = {
                        'element': menu_item,
                        'date': current_date if current_date else "Date Not Found",
                        'meal_type': current_meal if current_meal else "Meal Not Found",
                        'onclick': menu_item.get_attribute('onclick')
                    }
                    structured_meals.append(meal_info)
            
            print(f"\nStructured {len(structured_meals)} meal periods")
            for meal in structured_meals:
                print(f"  {meal['date']} - {meal['meal_type']}")
            
            return structured_meals
            
        except Exception as e:
            print(f"Error getting meal structure: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
    
    def click_meal(self, meal_element):
        """Click on a specific meal to load its items"""
        try:
            print("Clicking meal...")
            
            # Try to get the onclick attribute and execute it directly
            onclick = meal_element.get_attribute('onclick')
            if onclick:
                print(f"  Using onclick: {onclick}")
                self.driver.execute_script(onclick)
            else:
                # Fallback to clicking the element
                self.driver.execute_script("arguments[0].click();", meal_element)
            
            time.sleep(8)
            print("Meal loaded")
            return True
        except Exception as e:
            print(f"Error clicking meal: {str(e)}")
            return False
    
    def extract_nutrition_info(self, max_items=None):
        """Extract nutrition information for each menu item by clicking on them"""
        try:
            print("Extracting nutrition information...")
            
            items_data = []
            items = self.driver.find_elements(By.CSS_SELECTOR, "a.cbo_nn_itemHover")
            print(f"Found {len(items)} clickable items")
            
            if len(items) == 0:
                print("No clickable items found!")
                return items_data
            
            # Limit items for testing if max_items is specified
            if max_items:
                items_to_process = items[:max_items]
                print(f"Processing first {len(items_to_process)} items (testing mode)\n")
            else:
                items_to_process = items
                print(f"Processing all {len(items_to_process)} items\n")
            
            for i, item in enumerate(items_to_process, 1):
                try:
                    # Extract food name
                    food_name = ""
                    try:
                        food_name = item.text.strip()
                    except:
                        pass
                    
                    if not food_name:
                        try:
                            inner_html = item.get_attribute('innerHTML')
                            soup = BeautifulSoup(inner_html, 'html.parser')
                            food_name = soup.get_text().strip()
                        except:
                            pass
                    
                    print(f"  {i}. {food_name}")
                    
                    if not food_name:
                        continue
                    
                    # Click to open nutrition modal
                    print(f"     → Clicking...")
                    try:
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", item)
                        time.sleep(0.5)
                        self.driver.execute_script("arguments[0].click();", item)
                    except:
                        item.click()
                    
                    time.sleep(4)
                    
                    # Extract nutrition info
                    nutrition_info = self.extract_nutrition_from_modal(food_name)
                    items_data.append(nutrition_info)
                    
                    if nutrition_info.get('nutrition'):
                        print(f"     ✓ Extracted {len(nutrition_info['nutrition'])} nutrition fields")
                    else:
                        print(f"     ✗ No nutrition data found")
                    
                    # Close modal
                    self.close_modal()
                    time.sleep(1)
                
                except Exception as e:
                    print(f"    ERROR: {str(e)}")
                    continue
            
            print(f"\n✓ Extracted nutrition info for {len(items_data)} items")
            return items_data
            
        except Exception as e:
            print(f"Error: {str(e)}")
            return []
    
    def close_modal(self):
        """Close any open modal"""
        try:
            # Try various methods to close modal
            close_selectors = [
                "button[class*='close']",
                "button[aria-label*='close']",
                ".modal button.close",
                "button[data-dismiss='modal']"
            ]
            
            for selector in close_selectors:
                try:
                    close_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    self.driver.execute_script("arguments[0].click();", close_button)
                    return
                except:
                    pass
            
            # Try Escape key
            try:
                self.driver.execute_script("document.dispatchEvent(new KeyboardEvent('keydown', {key: 'Escape'}))")
            except:
                pass
                
        except Exception as e:
            print(f"     Error closing modal: {str(e)}")
    
    def extract_nutrition_from_modal(self, food_name):
        """Extract nutrition information from the modal"""
        nutrition_info = {
            'name': food_name,
            'serving_size': None,
            'nutrition': {}
        }
        
        try:
            # Find modal
            modal_body = None
            modal_selectors = [
                "div[class*='modal'][class*='show']",
                "div[role='dialog']",
                "div[class*='popup']"
            ]
            
            for selector in modal_selectors:
                try:
                    modal_body = self.driver.find_element(By.CSS_SELECTOR, selector)
                    break
                except:
                    pass
            
            if not modal_body:
                modal_body = self.driver.find_element(By.TAG_NAME, "body")
            
            modal_text = modal_body.text
            
            if not modal_text or len(modal_text) < 20:
                return nutrition_info
            
            lines = modal_text.split('\n')
            
            # Nutrition fields to extract
            nutrition_keywords = {
                'calories': ['calories'],
                'total_fat': ['total fat'],
                'saturated_fat': ['saturated fat'],
                'trans_fat': ['trans fat'],
                'cholesterol': ['cholesterol'],
                'sodium': ['sodium'],
                'potassium': ['potassium'],
                'total_carbohydrate': ['total carbohydrate', 'carbohydrate'],
                'dietary_fiber': ['dietary fiber'],
                'sugars': ['sugars'],
                'protein': ['protein']
            }
            
            for line in lines:
                line_lower = line.lower().strip()
                
                # Parse serving size
                if 'serving size' in line_lower:
                    nutrition_info['serving_size'] = line.split(':', 1)[-1].strip() if ':' in line else line
                
                # Parse nutrition values
                for key, keywords in nutrition_keywords.items():
                    for keyword in keywords:
                        if keyword in line_lower and key not in nutrition_info['nutrition']:
                            value = self.extract_nutrition_value(line, keyword)
                            if value:
                                nutrition_info['nutrition'][key] = value
                            break
            
        except Exception as e:
            print(f"       ERROR extracting nutrition: {str(e)}")
        
        return nutrition_info
    
    def extract_nutrition_value(self, line, keyword):
        """Extract nutrition value from a line"""
        try:
            keyword_pos = line.lower().find(keyword)
            if keyword_pos == -1:
                return None
            
            after_keyword = line[keyword_pos + len(keyword):].strip()
            
            value = ""
            for char in after_keyword:
                if char.isspace() and value:
                    break
                if char == '%':
                    break
                value += char
            
            value = value.strip()
            return value if value else None
            
        except:
            return None
    
    def scrape_all_with_complete_data(self):
        """Scrape all dining halls with nutrition info, meal periods, and entrée identification"""
        all_results = []
        
        print("="*80)
        print("Illinois Dining Complete Scraper - With Entrée Detection")
        print("="*80)
        
        dining_halls = self.scrape_dining_structure()
        
        if not dining_halls:
            print("Failed to get dining structure")
            return all_results
        
        # Testing mode limitations
        if self.testing_mode:
            print("\n[TESTING MODE] Limiting to first dining hall and first service")
            dining_halls = dining_halls[:1]  # Only first dining hall
            if dining_halls and dining_halls[0]['dining_services']:
                dining_halls[0]['dining_services'] = dining_halls[0]['dining_services'][:1]  # Only first service
        
        for hall in dining_halls:
            hall_name = hall['dining_hall']
            
            print(f"\n{'='*80}")
            print(f"Processing: {hall_name}")
            print(f"{'='*80}")
            
            if not hall['dining_services']:
                continue
            
            for service in hall['dining_services']:
                service_name = service['service_name']
                service_id = service['service_id']
                
                if not self.navigate_to_service(service_id, service_name):
                    continue
                
                # Get all meals
                structured_meals = self.get_all_meals_structured()
                
                if not structured_meals:
                    continue
                
                # Testing mode: limit number of meals
                if self.testing_mode:
                    print(f"[TESTING MODE] Limiting to first meal period only")
                    structured_meals = structured_meals[:1]  # Only process first meal
                
                for meal_idx, meal_info in enumerate(structured_meals):
                    if not meal_info['element']:
                        continue
                    
                    print(f"\n[Meal {meal_idx + 1}/{len(structured_meals)}]")
                    print(f"Date: {meal_info['date']}, Meal: {meal_info['meal_type']}")
                    
                    if not self.click_meal(meal_info['element']):
                        continue
                    
                    # Extract nutrition info
                    nutrition_items = self.extract_nutrition_info(max_items=self.max_items_per_meal)
                    
                    # Store results with meal info and entrée detection
                    for item_data in nutrition_items:
                        # Determine if item is an entrée
                        is_entree = self.determine_if_entree(
                            item_data['name'],
                            item_data.get('serving_size'),
                            item_data.get('nutrition', {}).get('calories')
                        )
                        
                        result = {
                            'dining_hall': hall_name,
                            'service': service_name,
                            'date': meal_info['date'],
                            'meal_type': meal_info['meal_type'],
                            'is_entree': is_entree,  # New column
                            'name': item_data['name'],
                            'serving_size': item_data.get('serving_size'),
                            'calories': item_data.get('nutrition', {}).get('calories', 'N/A'),
                            'total_fat': item_data.get('nutrition', {}).get('total_fat', 'N/A'),
                            'saturated_fat': item_data.get('nutrition', {}).get('saturated_fat', 'N/A'),
                            'trans_fat': item_data.get('nutrition', {}).get('trans_fat', 'N/A'),
                            'cholesterol': item_data.get('nutrition', {}).get('cholesterol', 'N/A'),
                            'sodium': item_data.get('nutrition', {}).get('sodium', 'N/A'),
                            'potassium': item_data.get('nutrition', {}).get('potassium', 'N/A'),
                            'total_carbohydrate': item_data.get('nutrition', {}).get('total_carbohydrate', 'N/A'),
                            'dietary_fiber': item_data.get('nutrition', {}).get('dietary_fiber', 'N/A'),
                            'sugars': item_data.get('nutrition', {}).get('sugars', 'N/A'),
                            'protein': item_data.get('nutrition', {}).get('protein', 'N/A')
                        }
                        all_results.append(result)
                    
                    print(f"Stored nutrition for {len(nutrition_items)} items")
                    entree_count = sum(1 for r in all_results[-len(nutrition_items):] if r['is_entree'])
                    print(f"  - {entree_count} entrées, {len(nutrition_items) - entree_count} non-entrées")
                    
                    # Navigate back for next meal
                    if meal_idx < len(structured_meals) - 1:
                        print("Navigating back...")
                        if not self.navigate_to_service(service_id, service_name):
                            break
        
        print(f"\n{'='*80}")
        print("Complete scraping finished!")
        print(f"{'='*80}")
        print(f"Total items scraped: {len(all_results)}")
        entree_total = sum(1 for r in all_results if r['is_entree'])
        print(f"Total entrées: {entree_total}")
        print(f"Total non-entrées: {len(all_results) - entree_total}")
        
        return all_results
    
    def export_to_excel(self, all_results, filename=None):
        """Export results to Excel with complete data"""
        if not all_results:
            print("No data to export")
            return None
        
        try:
            if not filename:
                filename = f"complete_dining_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create DataFrame
            df = pd.DataFrame(all_results)
            
            # For now, ignore date and meal_type in the exported table
            # Sort by dining_hall, service, is_entree, name
            df = df.sort_values(['dining_hall', 'service', 'is_entree', 'name'],
                                ascending=[True, True, False, True])

            # Reorder columns (exclude date and meal_type for this export)
            column_order = [
                'dining_hall', 'service', 'is_entree', 'name', 'serving_size',
                'calories', 'total_fat', 'saturated_fat', 'trans_fat', 'cholesterol',
                'sodium', 'potassium', 'total_carbohydrate', 'dietary_fiber', 'sugars', 'protein'
            ]
            # Only keep columns that actually exist in df to avoid KeyError
            column_order = [c for c in column_order if c in df.columns]
            df = df[column_order]
            
            # Write to Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Complete Data', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Complete Data']
                
                # Auto-adjust columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format headers
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Highlight entrées with light background color
                entree_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                    # Check if this row is an entrée (column E is is_entree)
                    if row[4].value:  # is_entree column (index 4)
                        for cell in row:
                            cell.fill = entree_fill
            
            print(f"Exported to {filename}")
            print(f"Total rows: {len(df)}")
            
            # Summary statistics
            print("\nData Summary:")
            print(f"  Unique dining halls: {df['dining_hall'].nunique()}")
            print(f"  Unique dates: {df['date'].nunique()}")
            print(f"  Meal types: {df['meal_type'].value_counts().to_dict()}")
            print(f"  Entrées: {df['is_entree'].sum()}")
            print(f"  Non-entrées: {(~df['is_entree']).sum()}")
            
            return filename
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def close(self):
        """Close the browser"""
        self.driver.quit()


if __name__ == "__main__":
    # Set testing_mode=False for full scraping
    TESTING_MODE = False
    
    scraper = NutritionScraperComplete(testing_mode=TESTING_MODE)
    
    try:
        print("\n" + "="*80)
        if TESTING_MODE:
            print("RUNNING IN TESTING MODE")
            print("- Will scrape only 5 items per meal")
            print("- Consider limiting dining halls and meals for faster testing")
        print("="*80 + "\n")
        all_results = scraper.scrape_all_with_complete_data()
        
        if all_results:
            print(f"\n{'='*80}")
            print("Final Results Summary")
            print(f"{'='*80}")
            
            # Summary statistics
            unique_halls = set(r['dining_hall'] for r in all_results)
            unique_dates = set(r['date'] for r in all_results if r['date'])
            unique_meals = set(r['meal_type'] for r in all_results if r['meal_type'])
            entree_count = sum(1 for r in all_results if r['is_entree'])
            
            print(f"Total items: {len(all_results)}")
            print(f"Dining halls: {len(unique_halls)}")
            print(f"Dates covered: {', '.join(sorted(unique_dates))}")
            print(f"Meal types: {', '.join(sorted(unique_meals))}")
            print(f"Entrées: {entree_count} ({entree_count*100//len(all_results)}%)")
            print(f"Non-entrées: {len(all_results) - entree_count} ({(len(all_results)-entree_count)*100//len(all_results)}%)")
            
            # Export
            print(f"\nExporting complete data...")
            excel_file = scraper.export_to_excel(all_results)
            
            if excel_file:
                print(f"\n✓ Success! Excel file: {excel_file}")
            
            # Sample output
            print(f"\nSample entrées:")
            entrees = [r for r in all_results if r['is_entree']][:3]
            for r in entrees:
                print(f"  • {r['name']} ({r['meal_type']}) - {r['calories']} cal")
            
            print(f"\nSample non-entrées:")
            non_entrees = [r for r in all_results if not r['is_entree']][:3]
            for r in non_entrees:
                print(f"  • {r['name']} ({r['meal_type']}) - {r['calories']} cal")
            
        else:
            print("No data scraped")
    
    except KeyboardInterrupt:
        print("\n\nInterrupted by user")
    except Exception as e:
        print(f"\nError: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        scraper.close()
        print("Browser closed.")